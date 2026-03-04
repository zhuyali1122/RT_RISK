"""
RT_RISK Web 应用 - 角色化资产管理平台
"""
import json
import os
from functools import wraps
from flask import (
    Flask, render_template, jsonify, request,
    session, redirect, url_for, send_from_directory, send_file
)
from werkzeug.utils import secure_filename

app = Flask(__name__)
# Vercel 等 serverless 需固定 secret_key，否则 session 无法跨请求
app.secret_key = os.getenv("SECRET_KEY", "dev-secret-change-me-in-production")
# 子路径部署，如 chuanx.xyz/rtrisk，设置 APP_ROOT=/rtrisk
APP_ROOT = (os.getenv("APP_ROOT") or "").rstrip("/")

BASE_DIR = os.path.dirname(__file__)
CONFIG_PATH = os.path.join(BASE_DIR, "config", "users.json")
PORTFOLIO_PATH = os.path.join(BASE_DIR, "config", "portfolio.json")
DD_CHECKLIST_PATH = os.path.join(BASE_DIR, "config", "dd_checklist.json")
TRANSACTIONS_PATH = os.path.join(BASE_DIR, "config", "transactions.json")
VINTAGE_PORTFOLIO_PATH = os.path.join(BASE_DIR, "config", "vintage_portfolio.json")
DPD_PORTFOLIO_PATH = os.path.join(BASE_DIR, "config", "dpd_portfolio.json")
MATURITY_PORTFOLIO_PATH = os.path.join(BASE_DIR, "config", "maturity_portfolio.json")
LOAN_DETAILS_PATH = os.path.join(BASE_DIR, "config", "loan_details.json")
PRODUCERS_PATH = os.path.join(BASE_DIR, "config", "producers.json")
DD_TEMPLATES_DIR = os.path.join(BASE_DIR, "dd_templates")
# Vercel serverless 仅 /tmp 可写
UPLOAD_DIR = os.path.join("/tmp", "rt_risk_uploads") if os.getenv("VERCEL") else os.path.join(BASE_DIR, "uploads")

try:
    os.makedirs(DD_TEMPLATES_DIR, exist_ok=True)
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    # 确保缓存目录存在（config/cache 在 .gitignore，部署后需创建）
    try:
        from kn_producer_cache import _ensure_cache_dir
        _ensure_cache_dir()
    except Exception:
        pass
except OSError:
    pass


@app.before_request
def _set_script_root():
    if APP_ROOT:
        request.environ["SCRIPT_NAME"] = APP_ROOT
    # Language: ?lang=en or ?lang=zh
    if request.args.get("lang") in ("en", "zh"):
        session["lang"] = request.args.get("lang")


def _get_lang():
    """Current UI language: en or zh (default)."""
    return session.get("lang", "zh")


@app.context_processor
def _inject_app_root():
    from translations import get_translations
    lang = _get_lang()
    trans = get_translations(lang)

    def t(key, default=""):
        return trans.get(key, default or key)

    cache_meta = None
    try:
        from kn_producer_cache import load_cache_meta
        raw = load_cache_meta()
        if raw:
            lu = raw.get("last_updated") or ""
            cache_meta = {
                "last_updated": lu,
                "last_updated_fmt": lu[:19].replace("T", " ") if lu else "",
                "system_cutover_date": raw.get("system_cutover_date") or "",
            }
    except Exception:
        pass

    return {
        "app_root": APP_ROOT,
        "lang": lang,
        "t": t,
        "translations_json": trans,
        "cache_meta": cache_meta,
    }


def load_user_config():
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def load_portfolio_data():
    """加载投资组合数据，无数据时返回空结构"""
    default = {
        "fund": {},
        "allocation_by_product": [],
        "allocation_by_region": [],
        "allocation_by_platform": [],
        "monthly_returns": [],
        "risk_metrics": {},
    }
    if not os.path.exists(PORTFOLIO_PATH):
        return default
    try:
        with open(PORTFOLIO_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
        return {**default, **data}
    except Exception:
        return default


def load_dd_checklist():
    with open(DD_CHECKLIST_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def load_partners():
    """从 producers.json 读取，返回与旧 partners 兼容的结构：assignments + partners"""
    try:
        from flask import g
        if hasattr(g, "_rt_load_partners"):
            return g._rt_load_partners
    except RuntimeError:
        pass
    producers = load_producers(json_only=True)
    if not producers:
        return {"assignments": {}, "partners": {}}
    active = {pid: p for pid, p in producers.items() if str(p.get("status", "active")).lower() in ("active", "")}
    producer_ids = list(active.keys())
    # 有 manage_partners 权限的用户可访问所有生产商
    assignments = {}
    try:
        users_cfg = load_user_config()
        roles = users_cfg.get("roles", {})
        for u in users_cfg.get("users", []):
            role = u.get("role", "")
            perms = roles.get(role, {}).get("permissions", [])
            if role in ("admin", "project_manager") or "manage_partners" in perms:
                assignments[u.get("username", "")] = producer_ids
    except Exception:
        assignments = {"pm": producer_ids, "admin": producer_ids}
    # 转为 partners 格式：region -> country
    partners = {}
    for pid, p in active.items():
        partners[pid] = {
            "id": p.get("id", pid),
            "name": p.get("name", pid),
            "country": p.get("region", p.get("country", "-")),
            "product_type": p.get("product_type", "-"),
            "status": p.get("status", "active"),
            "onboard_date": p.get("onboard_date", "-"),
            "contact": p.get("contact", "-"),
            "local_currency": p.get("currency", "USD"),
            "exchange_rate": p.get("exchange_rate", 1),
            "revenue_data": p.get("revenue_data", []),
            "risk_data": p.get("risk_data", []),
            "alerts": p.get("alerts", []),
            "priority_indicators": p.get("priority_indicators", {}),
        }
    out = {"assignments": assignments, "partners": partners}
    try:
        from flask import g
        g._rt_load_partners = out
    except (RuntimeError, Exception):
        pass
    return out


def load_transactions():
    with open(TRANSACTIONS_PATH, "r", encoding="utf-8") as f:
        return json.load(f)


def load_vintage_portfolio():
    if os.path.exists(VINTAGE_PORTFOLIO_PATH):
        with open(VINTAGE_PORTFOLIO_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def load_dpd_portfolio():
    if os.path.exists(DPD_PORTFOLIO_PATH):
        with open(DPD_PORTFOLIO_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def load_maturity_portfolio():
    if os.path.exists(MATURITY_PORTFOLIO_PATH):
        with open(MATURITY_PORTFOLIO_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def load_loan_details():
    if os.path.exists(LOAN_DETAILS_PATH):
        with open(LOAN_DETAILS_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def load_producers(json_only=False):
    """加载生产商配置：优先从 spv_config 表读取，否则 fallback 到 config/producers.json"""
    if json_only:
        try:
            from flask import g
            if hasattr(g, "_rt_producers_json"):
                return g._rt_producers_json
        except RuntimeError:
            pass
    from spv_config import load_producers_from_spv_config
    out = load_producers_from_spv_config(json_only=json_only)
    if json_only:
        try:
            from flask import g
            g._rt_producers_json = out
        except (RuntimeError, Exception):
            pass
    return out


def _find_loan(partner_id, loan_id):
    """Find loan in vintage, dpd, or maturity portfolio."""
    v = load_vintage_portfolio()
    for month, loans in v.get(partner_id, {}).items():
        for loan in loans:
            if loan.get("loan_id") == loan_id:
                return loan
    d = load_dpd_portfolio()
    for bucket, loans in d.get(partner_id, {}).items():
        for loan in loans:
            if loan.get("loan_id") == loan_id:
                return loan
    m = load_maturity_portfolio()
    for month, loans in m.get(partner_id, {}).items():
        for loan in loans:
            if loan.get("loan_id") == loan_id:
                return loan
    return None


def _portfolio_stats(loans):
    """Compute customer_type, product_type stats and key KPIs for portfolio."""
    n = len(loans)
    if n == 0:
        return {
            "customer_stats": {"new_count": 0, "returning_count": 0, "new_pct": 0, "returning_pct": 0},
            "product_stats": [],
            "credit_ratings": [],
            "credit_stats": [],
            "customer_pie_gradient": "#E2E8F0 0% 100%",
            "product_pie_gradient": "#E2E8F0 0% 100%",
            "credit_pie_gradient": "#E2E8F0 0% 100%",
            "kpi_stats": {
                "total_disbursement": 0, "avg_term": 0, "total_overdue": 0,
                "outstanding_balance": 0, "dpd7_pct": 0, "mob1_pct": None,
            },
        }
    new_count = sum(1 for l in loans if l.get("customer_type") == "new")
    ret_count = n - new_count
    new_pct = round(new_count / n * 100, 1)
    ret_pct = round(ret_count / n * 100, 1)
    customer_pie = f"#2E7D6E 0% {new_pct}%, #4DB6AC {new_pct}% 100%"
    product_counts = {}
    for l in loans:
        pt = l.get("product_type") or "其他"
        product_counts[pt] = product_counts.get(pt, 0) + 1
    colors = ["#2E7D6E", "#4DB6AC", "#E8A838", "#3B4A6A", "#D8434F"]
    product_stats = []
    cum = 0
    pie_parts = []
    for i, (name, cnt) in enumerate(product_counts.items()):
        pct = round(cnt / n * 100, 1)
        product_stats.append({"name": name, "count": cnt, "pct": pct, "color": colors[i % len(colors)]})
        pie_parts.append(f"{colors[i % len(colors)]} {cum}% {cum + pct}%")
        cum += pct
    product_pie = ", ".join(pie_parts) if pie_parts else "#E2E8F0 0% 100%"

    credit_counts = {}
    for l in loans:
        cr = l.get("credit_rating") or "-"
        credit_counts[cr] = credit_counts.get(cr, 0) + 1
    credit_ratings = sorted(credit_counts.keys())
    credit_stats = []
    cum = 0
    credit_pie_parts = []
    cr_colors = ["#2E7D6E", "#4DB6AC", "#E8A838", "#3B4A6A", "#D8434F"]
    for i, cr in enumerate(credit_ratings):
        cnt = credit_counts[cr]
        pct = round(cnt / n * 100, 1)
        credit_stats.append({"name": cr, "count": cnt, "pct": pct, "color": cr_colors[i % len(cr_colors)]})
        credit_pie_parts.append(f"{cr_colors[i % len(cr_colors)]} {cum}% {cum + pct}%")
        cum += pct
    credit_pie = ", ".join(credit_pie_parts) if credit_pie_parts else "#E2E8F0 0% 100%"

    # Key KPIs
    total_disb = sum(float(l.get("disbursement_amount") or 0) for l in loans)
    terms = [float(l.get("term_month") or 0) for l in loans]
    avg_term = round(sum(terms) / n, 1) if terms else 0
    total_overdue = sum(
        float(l.get("overdue_principal") or 0) + float(l.get("overdue_interest") or 0) + float(l.get("overdue_penalty") or 0)
        for l in loans
    )
    outstanding = sum(float(l.get("outstanding_principal") or 0) for l in loans)
    dpd7_count = sum(1 for l in loans if (l.get("dpd") or 0) >= 7)
    dpd7_pct = round(dpd7_count / n * 100, 2) if n else 0
    mob1_rates = [float(l.get("mob1_rate") or 0) for l in loans if l.get("mob1_rate") is not None]
    mob1_pct = round(sum(mob1_rates) / len(mob1_rates) * 100, 2) if mob1_rates else None

    return {
        "customer_stats": {"new_count": new_count, "returning_count": ret_count, "new_pct": new_pct, "returning_pct": ret_pct},
        "product_stats": product_stats,
        "credit_ratings": credit_ratings,
        "credit_stats": credit_stats,
        "customer_pie_gradient": customer_pie,
        "product_pie_gradient": product_pie,
        "credit_pie_gradient": credit_pie,
        "kpi_stats": {
            "total_disbursement": total_disb,
            "avg_term": avg_term,
            "total_overdue": total_overdue,
            "outstanding_balance": outstanding,
            "dpd7_pct": dpd7_pct,
            "mob1_pct": mob1_pct,
        },
    }


def fmt_usd(n):
    if n >= 1e6:
        return f"${n/1e6:.1f}M"
    if n >= 1e3:
        return f"${n/1e3:.0f}K"
    return f"${n:,.0f}"


def _fmt_local(n):
    """本币格式（无数额符号）"""
    if n is None or (isinstance(n, float) and (n != n or n < 0)):
        return "-"
    try:
        x = float(n)
    except (TypeError, ValueError):
        return "-"
    if x >= 1e6:
        return f"{x/1e6:.1f}M"
    if x >= 1e3:
        return f"{x/1e3:.0f}K"
    return f"{x:,.0f}"


def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if "user" not in session:
            # API 请求返回 JSON，避免 fetch 解析 HTML 报错
            if request.path.startswith("/api/") or "application/json" in (request.headers.get("Accept") or ""):
                return jsonify({"error": "请先登录"}), 401
            return redirect(url_for("login_page"))
        return f(*args, **kwargs)
    return decorated


@app.route("/")
def index():
    if "user" in session:
        return redirect(url_for("dashboard"))
    return redirect(url_for("login_page"))


@app.route("/login")
def login_page():
    if "user" in session:
        return redirect(url_for("dashboard"))
    return render_template("login.html")


@app.route("/api/login", methods=["POST"])
def api_login():
    data = request.get_json() or {}
    username = data.get("username", "").strip()
    password = data.get("password", "")

    # Set lang from login request if provided
    if data.get("lang") in ("en", "zh"):
        session["lang"] = data["lang"]
    lang = _get_lang()

    config = load_user_config()

    for user in config["users"]:
        if user["username"] == username and user["password"] == password:
            role_info = config["roles"].get(user["role"], {})
            role_label = role_info.get("label_en" if lang == "en" else "label", user["role"])
            session["user"] = {
                "username": user["username"],
                "display_name": user["display_name"],
                "email": user["email"],
                "role": user["role"],
                "role_label": role_label,
                "permissions": role_info.get("permissions", []),
            }

            # 仅 Admin 登录时后台刷新缓存；PM/Investor 仅读缓存，由 Admin 通过 Dashboard「管理」按钮刷新
            if user["role"] == "admin":
                try:
                    from kn_producer_cache import refresh_producer_full_cache_async
                    refresh_producer_full_cache_async()
                except Exception:
                    pass

            redirect_url = (APP_ROOT or "") + "/dashboard"
            return jsonify({"success": True, "redirect": redirect_url})

    err_msg = "Invalid username or password" if lang == "en" else "用户名或密码错误"
    return jsonify({"success": False, "message": err_msg})


def _user_with_role_label(user, lang):
    """Ensure user.role_label matches current lang."""
    cfg = load_user_config()
    role_info = cfg.get("roles", {}).get(user.get("role", ""), {})
    label = role_info.get("label_en" if lang == "en" else "label", user.get("role", ""))
    return {**user, "role_label": label}


@app.route("/dashboard")
@login_required
def dashboard():
    user = _user_with_role_label(session["user"], _get_lang())
    return render_template("dashboard.html", user=user)


@app.route("/admin/cache-refresh")
@login_required
def admin_cache_refresh():
    """Admin 数据刷新页面：显示更新时间、系统切日，需确认后刷新，可查看日志"""
    user = _user_with_role_label(session["user"], _get_lang())
    if not _is_admin(user):
        return redirect(url_for("dashboard"))
    cache_meta = None
    try:
        from kn_producer_cache import load_cache_meta, load_refresh_log
        raw = load_cache_meta()
        if raw:
            lu = raw.get("last_updated") or ""
            cache_meta = {
                "last_updated": lu,
                "last_updated_fmt": lu[:19].replace("T", " ") if lu else "",
                "system_cutover_date": raw.get("system_cutover_date") or "",
            }
        refresh_logs = load_refresh_log()
        refresh_logs_text = "".join(refresh_logs) if refresh_logs else ""
    except Exception:
        refresh_logs = []
        refresh_logs_text = ""
    return render_template(
        "admin_cache_refresh.html",
        user=user,
        cache_meta=cache_meta,
        refresh_logs=refresh_logs,
        refresh_logs_text=refresh_logs_text,
    )


# Help 文档 slug -> 文件名映射（避免 URL 编码问题）
HELP_DOC_MAP = {
    "metrics": "指标计算说明",
    "指标计算说明": "指标计算说明",
}


def _extract_toc_from_markdown(md_content):
    """从 Markdown 提取层级目录，返回 [(level, title, slug), ...]"""
    import re
    toc = []
    for m in re.finditer(r"^(#{1,3})\s+(.+)$", md_content, re.MULTILINE):
        level = len(m.group(1))
        title = m.group(2).strip()
        slug = re.sub(r"[^\w\u4e00-\u9fff\-]", "", title)[:30] or f"h{len(toc)}"
        toc.append({"level": level, "title": title, "slug": slug})
    return toc


@app.route("/help")
@app.route("/help/<path:doc_slug>")
def help_page(doc_slug=None):
    """Help 目录：指标计算说明等文档，无需登录即可访问"""
    if not doc_slug:
        doc_slug = "metrics"
    safe_slug = doc_slug.replace("..", "").strip().lower()
    if not safe_slug:
        safe_slug = "metrics"
    file_name = HELP_DOC_MAP.get(safe_slug) or HELP_DOC_MAP.get(doc_slug) or safe_slug
    help_dir = os.path.join(BASE_DIR, "docs", "help")
    md_path = os.path.join(help_dir, f"{file_name}.md")
    if not os.path.isfile(md_path):
        return render_template("help.html", user=session.get("user"), content="", toc=[], title="文档不存在", doc_slug=safe_slug, doc_list=[])
    try:
        with open(md_path, "r", encoding="utf-8") as f:
            md_content = f.read()
        toc = _extract_toc_from_markdown(md_content)
        try:
            import markdown
            html_content = markdown.markdown(md_content, extensions=["tables", "fenced_code"])
            html_content = _add_heading_ids(html_content, toc)
        except ImportError:
            html_content = f"<pre style='white-space:pre-wrap;font-size:0.9rem;'>{md_content}</pre><p style='color:#8A95A5;font-size:0.8rem;margin-top:16px;'>提示：安装 markdown 包可获得更好排版（pip install markdown）</p>"
            toc = _extract_toc_from_markdown(md_content)
    except Exception as e:
        html_content = f"<p>读取文档失败: {e}</p>"
        toc = []
    doc_list = [{"slug": k, "title": v} for k, v in HELP_DOC_MAP.items() if k.isascii()]
    if not doc_list:
        doc_list = [{"slug": "metrics", "title": "指标计算说明"}]
    user = session.get("user")
    return render_template(
        "help.html",
        user=user,
        content=html_content,
        toc=toc,
        title=file_name,
        doc_slug=safe_slug,
        doc_list=doc_list,
    )


def _add_heading_ids(html_content, toc):
    """为 HTML 中的 h1/h2/h3 添加 id 属性，便于锚点跳转"""
    import re
    idx = [0]
    def repl(m):
        level, rest = m.group(1), m.group(2)
        slug = toc[idx[0]]["slug"] if idx[0] < len(toc) else f"h{idx[0]}"
        idx[0] += 1
        return f'<h{level} id="{slug}">{rest}</h{level}>'
    return re.sub(r"<h([123])>([\s\S]*?)</h\1>", repl, html_content)


@app.route("/alert/panel")
@login_required
def alert_panel():
    user = session["user"]
    if "alert_panel" not in user.get("permissions", []):
        return redirect(url_for("dashboard"))
    data = load_partners()
    partners_list = []
    for pid, p in data["partners"].items():
        alerts = p.get("alerts", [])
        pi = p.get("priority_indicators", {})
        cov = pi.get("coverage_ratio", {})
        partners_list.append({
            "id": p["id"],
            "name": p["name"],
            "country": p["country"],
            "product_type": p["product_type"],
            "alerts": alerts,
            "coverage": cov,
        })
    return render_template(
        "alert_panel.html",
        user=user,
        partners=partners_list,
    )


def _portfolio_cumulative_stats():
    """
    从已投资平台（spv_internal_params）的 spv_id 访问 raw_loan，计算：
    - 累计放款总额：SUM(disbursement_amount)，按汇率转为 USD
    - 累计借款总量：COUNT(loan_id)
    - 累计借款人数：COUNT(DISTINCT customer_id)
    """
    try:
        from spv_internal_params import load_invested_spv_ids_for_portfolio
        from kn_risk_query import query_portfolio_cumulative_stats
        spv_ids = load_invested_spv_ids_for_portfolio()
        return query_portfolio_cumulative_stats(spv_ids)
    except Exception:
        return {
            "cumulative_disbursement": 0,
            "cumulative_loan_count": 0,
            "cumulative_borrower_count": 0,
        }


# 国家名映射：中文 -> ECharts 世界地图用英文
_COUNTRY_NAME_MAP = {
    "中国": "China", "中国内地": "China", "中国香港": "Hong Kong", "印度尼西亚": "Indonesia", "印尼": "Indonesia", "越南": "Vietnam", "泰国": "Thailand",
    "菲律宾": "Philippines", "马来西亚": "Malaysia", "新加坡": "Singapore",
    "印度": "India", "孟加拉国": "Bangladesh", "巴基斯坦": "Pakistan", "斯里兰卡": "Sri Lanka",
    "日本": "Japan", "韩国": "South Korea",
    "墨西哥": "Mexico", "巴西": "Brazil", "哥伦比亚": "Colombia", "阿根廷": "Argentina",
    "智利": "Chile", "秘鲁": "Peru", "尼日利亚": "Nigeria", "肯尼亚": "Kenya",
    "南非": "South Africa", "埃及": "Egypt", "摩洛哥": "Morocco",
    "美国": "United States", "英国": "United Kingdom", "德国": "Germany", "法国": "France",
}

def _aggregate_producer_data(records):
    """按地域、行业、成熟度、场景、类型、国家聚合"""
    from collections import defaultdict
    by_region = defaultdict(list)
    by_country = defaultdict(list)  # 国家级，用于世界地图
    by_industry = defaultdict(list)
    by_maturity = defaultdict(list)
    by_scenario = defaultdict(list)
    by_type = defaultdict(list)
    region_detail_map = defaultdict(set)
    for r in records:
        region = r.get("region") or r.get("地域") or "未分类"
        region_d = r.get("region_detail") or r.get("地区明细") or region
        industry = r.get("industry") or r.get("行业") or "未分类"
        maturity = r.get("maturity") or r.get("项目成熟程度") or "未分类"
        scenario = r.get("scenario") or r.get("场景") or "未分类"
        typ = r.get("type") or r.get("类型") or "未分类"
        name = r.get("name") or r.get("名称") or r.get("生产商") or "-"
        rec = {**r, "name": name, "region": region, "region_detail": region_d, "industry": industry, "maturity": maturity, "scenario": scenario, "type": typ}
        by_region[region].append(rec)
        country_en = _COUNTRY_NAME_MAP.get(region_d, region_d)
        by_country[country_en].append(rec)
        by_industry[industry].append(rec)
        by_maturity[maturity].append(rec)
        by_scenario[scenario].append(rec)
        by_type[typ].append(rec)
        region_detail_map[region].add(region_d)
    return {
        "by_region": {k: {"count": len(v), "items": v} for k, v in by_region.items()},
        "by_country": {k: {"count": len(v), "items": v} for k, v in by_country.items()},
        "by_industry": {k: {"count": len(v), "items": v} for k, v in by_industry.items()},
        "by_maturity": {k: {"count": len(v), "items": v} for k, v in by_maturity.items()},
        "by_scenario": {k: {"count": len(v), "items": v} for k, v in by_scenario.items()},
        "by_type": {k: {"count": len(v), "items": v} for k, v in by_type.items()},
        "region_details": {k: list(v) for k, v in region_detail_map.items()},
    }


@app.route("/api/producers/data")
@login_required
def api_producers_data():
    """生产商拓展列表数据（含聚合统计）"""
    if "view_portfolio" not in session.get("user", {}).get("permissions", []):
        return jsonify({"error": "权限不足"}), 403
    try:
        from feishu_producer import load_producer_data
        data = load_producer_data()
        records = data.get("records", [])
        agg = _aggregate_producer_data(records)
        return jsonify({
            "records": records,
            "aggregation": agg,
            "source": data.get("source", "unknown"),
            "total": len(records),
        })
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/portfolio")
@login_required
def portfolio():
    user = _user_with_role_label(session["user"], _get_lang())
    if "view_portfolio" not in user.get("permissions", []):
        return redirect(url_for("dashboard"))
    portfolio_data = load_portfolio_data()
    fund = portfolio_data.setdefault("fund", {})
    fund.setdefault("name", "")
    fund.setdefault("name_cn", "")
    fund.setdefault("currency", "USD")
    fund.setdefault("inception_date", "")
    fund.setdefault("nav_date", "")
    fund.setdefault("nav_per_unit")
    fund.setdefault("total_aum")
    fund.setdefault("total_units")
    fund.setdefault("ytd_return")
    fund.setdefault("annualized_return")
    fund.setdefault("sharpe_ratio")
    fund.setdefault("weighted_duration_days")
    cache_only = _cache_only_mode(user)
    # PM/Investor 从缓存读取累计指标与平台持仓；Admin 从 DB
    if cache_only:
        try:
            from kn_producer_cache import load_producer_full_cache
            cache_data, _ = load_producer_full_cache()
            cum = (cache_data or {}).get("portfolio_cumulative_stats") or {}
            allocation = (cache_data or {}).get("allocation_by_platform")
        except Exception:
            cum = {}
            allocation = None
        fund["cumulative_disbursement"] = cum.get("cumulative_disbursement", 0)
        fund["cumulative_loan_count"] = cum.get("cumulative_loan_count", 0)
        fund["cumulative_borrower_count"] = cum.get("cumulative_borrower_count", 0)
        if allocation:
            portfolio_data["allocation_by_platform"] = allocation
    else:
        cum = _portfolio_cumulative_stats()
        fund["cumulative_disbursement"] = cum["cumulative_disbursement"]
        fund["cumulative_loan_count"] = cum["cumulative_loan_count"]
        fund["cumulative_borrower_count"] = cum["cumulative_borrower_count"]
        try:
            from spv_internal_params import load_all_spv_internal_params_for_portfolio
            trades = load_all_spv_internal_params_for_portfolio()
            if trades:
                total_principal = sum(t.get("principal_amount") or 0 for t in trades)
                allocation = []
                for t in trades:
                    pct = (t.get("principal_amount") or 0) / total_principal if total_principal > 0 else 0
                    agreed = t.get("agreed_rate") or 0
                    agreed_pct = agreed * 100 if agreed <= 1 else agreed
                    allocation.append({
                        "name": t.get("name") or t.get("spv_id") or "-",
                        "value": t.get("principal_amount") or 0,
                        "pct": pct,
                        "type": t.get("product_type") or "-",
                        "region": t.get("region") or "-",
                        "principal_amount": t.get("principal_amount"),
                        "agreed_rate": agreed_pct,
                        "effective_date": t.get("effective_date") or "-",
                    })
                portfolio_data["allocation_by_platform"] = allocation
        except Exception:
            pass
    return render_template("portfolio.html", user=user, portfolio=portfolio_data)


@app.route("/partner/apply")
@login_required
def partner_apply():
    user = session["user"]
    if "apply_partner" not in user.get("permissions", []):
        return redirect(url_for("dashboard"))
    dd_checklist = load_dd_checklist()
    return render_template("partner_apply.html", user=user, dd_checklist=dd_checklist)


@app.route("/partner/manage")
@login_required
def partner_manage():
    user = _user_with_role_label(session["user"], _get_lang())
    if "manage_partners" not in user.get("permissions", []):
        return redirect(url_for("dashboard"))

    # 加载统一缓存
    cache_data = None
    full_cache = {}
    try:
        from kn_producer_cache import load_producer_full_cache
        cache_data, _ = load_producer_full_cache()
        full_cache = (cache_data or {}).get("producers", {})
    except Exception:
        pass
    cache_exists = bool(full_cache)
    cache_only = _cache_only_mode(user)

    # PM/Investor 仅读缓存：无缓存时提示联系管理员
    if cache_only and not cache_exists:
        return render_template("partner_list.html", user=user, partners=[], partners_revenue=[], partners_cashflow=[],
            data_source="cache", no_cache_hint=True, show_refresh=False)

    # 生产商列表：缓存存在时从缓存；否则 Admin 从 DB
    if cache_exists:
        producers_active = [
            {"id": sid, "name": p.get("name", sid), "region": p.get("region", "-"), "product_type": p.get("product_type", "-"),
             "status": p.get("status", "active"), "onboard_date": p.get("onboard_date", "-"), "exchange_rate": p.get("exchange_rate", 1),
             "currency": p.get("currency", "USD")}
            for sid, p in full_cache.items() if str(p.get("status", "active")).lower() in ("active", "")
        ]
    else:
        try:
            from project_loader import load_projects_with_internal_params
            projects = load_projects_with_internal_params()
        except Exception:
            projects = {}
        producers_active = [p for p in projects.values() if str(p.get("status", "active")).lower() in ("active", "")]
        if not producers_active:
            producers = load_producers()
            producers_active = [p for p in producers.values() if p.get("status") == "active"]

    partners = []
    partners_revenue = []
    partners_cashflow = []
    # 缓存存在时用 JSON 派生的映射，避免 DB
    if cache_exists:
        spv_map = {p["id"]: p["id"] for p in producers_active}
        if "kn" in spv_map:
            spv_map["partner_beta"] = "kn"
    else:
        spv_map = _get_partner_spv_map()
    for prod in producers_active:
        pid = prod["id"]
        spv_id = spv_map.get(pid) or pid
        pc = (full_cache or {}).get(spv_id) or (full_cache or {}).get(str(spv_id).strip().lower()) if full_cache else None

        partner_like = {
            "id": pid,
            "name": prod.get("name", pid),
            "country": prod.get("region", "-"),
            "product_type": prod.get("product_type", "-"),
            "status": prod.get("status", "active"),
            "onboard_date": prod.get("onboard_date", "-"),
            "monthly_volume": 0,
            "monthly_volume_fmt": "-",
        }
        if cache_exists:
            risk_data = (pc or {}).get("risk_data", [])
        else:
            risk_data = _load_risk_data_for_partner(pid, partner_like)
        # 缓存存在时用 prod 的配置，避免 _get_producer_config 访问 DB
        if cache_exists:
            rate = float(prod.get("exchange_rate", 1) or 1)
            local_currency = prod.get("currency", "USD") or "USD"
            env_key = f"{spv_id.upper()}_EXCHANGE_RATE"
            if os.getenv(env_key):
                try:
                    rate = float(os.getenv(env_key))
                except ValueError:
                    pass
        else:
            cfg = _get_producer_config(pid)
            rate = float((cfg.get("exchange_rate") if cfg else None) or prod.get("exchange_rate", 1) or 1)
            local_currency = (cfg.get("currency") if cfg else None) or prod.get("currency", "USD")
        is_usd = (local_currency or "USD") == "USD"

        if risk_data:
            latest = sorted(risk_data, key=lambda r: r.get("stat_date", ""), reverse=True)[0]
            od1 = float(latest.get("overdue_1_plus_ratio", 0))
            m0 = float(latest.get("m0_ratio", 0))
            cb = float(latest.get("current_balance", 0))
            cd = float(latest.get("cumulative_disbursement", 0))
            usd_cb = cb / rate if rate else cb
            usd_cd = cd / rate if rate else cd
            partner_like["latest"] = {
                "stat_date": latest["stat_date"],
                "current_balance_fmt": fmt_usd(usd_cb),
                "cum_disb_fmt": fmt_usd(usd_cd),
                "current_balance_local": _fmt_local(cb),
                "current_balance_usd": fmt_usd(usd_cb),
                "cum_disb_local": _fmt_local(cd),
                "cum_disb_usd": fmt_usd(usd_cd),
                "m0_ratio_fmt": f"{m0*100:.2f}%",
                "m0_class": "good" if m0 >= 0.96 else "warn" if m0 >= 0.93 else "danger",
                "od1_fmt": f"{od1*100:.2f}%",
                "od1_class": "good" if od1 < 0.03 else "warn" if od1 < 0.05 else "danger",
                "active_borrowers": f"{int(latest.get('active_borrowers', 0)):,}",
            }
        else:
            partner_like["latest"] = {
                "stat_date": "-",
                "current_balance_fmt": "-",
                "cum_disb_fmt": "-",
                "current_balance_local": "-",
                "current_balance_usd": "-",
                "cum_disb_local": "-",
                "cum_disb_usd": "-",
                "m0_ratio_fmt": "-",
                "m0_class": "",
                "od1_fmt": "-",
                "od1_class": "",
                "active_borrowers": "-",
            }
        partner_like["local_currency"] = local_currency
        partner_like["exchange_rate"] = rate
        partner_like["is_usd"] = is_usd

        # 收益数据：缓存存在时仅用统一缓存；否则回退到单独缓存或 prod
        if cache_exists:
            rev_data = (pc or {}).get("revenue_data", [])
        else:
            rev_data = prod.get("revenue_data", [])
            if not rev_data:
                try:
                    from kn_revenue_cache import load_revenue_cache
                    cached_rev, _ = load_revenue_cache(pid)
                    if cached_rev:
                        rev_data = cached_rev
                except Exception:
                    pass

        # 月放贷规模：从 revenue_data 最新月 disbursement
        monthly_vol = 0
        if rev_data:
            monthly_vol = float(rev_data[-1].get("disbursement", 0) or 0)
        usd_mv = monthly_vol / rate if rate else monthly_vol
        partner_like["monthly_volume"] = monthly_vol
        partner_like["monthly_volume_local"] = _fmt_local(monthly_vol) if monthly_vol else "-"
        partner_like["monthly_volume_usd"] = fmt_usd(usd_mv) if monthly_vol else "-"
        partner_like["monthly_volume_fmt"] = fmt_usd(usd_mv) if monthly_vol else "-"
        partners.append(partner_like)
        if cache_exists:
            rev_local_currency = local_currency
            rev_exchange_rate = rate
        else:
            cfg = _get_producer_config(pid)
            rev_local_currency = (cfg.get("currency") if cfg else None) or prod.get("currency", "USD")
            rev_exchange_rate = float((cfg.get("exchange_rate") if cfg else None) or prod.get("exchange_rate", 1) or 1)
        partners_revenue.append({
            "id": pid,
            "name": partner_like["name"],
            "country": partner_like["country"],
            "product_type": partner_like["product_type"],
            "revenue_data": rev_data,
            "local_currency": rev_local_currency,
            "exchange_rate": rev_exchange_rate,
        })
        # 现金流预测：缓存存在时仅用统一缓存；否则回退到单独缓存或 DB
        if cache_exists:
            cashflow_data = (pc or {}).get("cashflow_data", [])
        else:
            cashflow_data = []
            try:
                from kn_cashflow_cache import load_cashflow_cache
                cached_cf, _, _ = load_cashflow_cache(pid)
                if cached_cf:
                    cashflow_data = cached_cf
                else:
                    from kn_cashflow import compute_cashflow_forecast
                    coll_rate = 0.98
                    if rev_data:
                        coll_rate = rev_data[-1].get("collection_rate", 0.98) or 0.98
                    cf = compute_cashflow_forecast(spv_id=pid, months_ahead=12, collection_rate=coll_rate)
                    cashflow_data = cf.get("forecast", [])
            except Exception:
                pass
        partners_cashflow.append({
            "id": pid,
            "name": partner_like["name"],
            "country": partner_like["country"],
            "product_type": partner_like["product_type"],
            "cashflow_data": cashflow_data,
            "total_expected": sum(r.get("expected_inflow", 0) for r in cashflow_data),
            "local_currency": rev_local_currency,
            "exchange_rate": rev_exchange_rate,
        })
    return render_template(
        "partner_list.html", user=user, partners=partners,
        partners_revenue=partners_revenue,
        partners_cashflow=partners_cashflow,
        data_source="cache" if cache_exists else "database",
        show_refresh=not cache_only,
    )


def _get_producer_data_from_full_cache(spv_id):
    """
    从统一缓存获取单个生产商数据
    返回: (pc, last_updated, cache_exists)
    - cache_exists=True: 缓存文件存在，必须仅用此数据，不再访问 DB 或单独缓存
    - cache_exists=False: 无统一缓存，可回退到单独缓存或 DB
    """
    try:
        from kn_producer_cache import load_producer_full_cache
        data, last_updated = load_producer_full_cache()
        cache_exists = bool(data and data.get("producers"))
        if not cache_exists:
            return None, None, False
        producers = data.get("producers", {})
        sid = str(spv_id or "").strip().lower()
        pc = producers.get(spv_id) or producers.get(sid)
        return pc, last_updated, True
    except Exception:
        return None, None, False


def _is_admin(user=None):
    """当前用户是否为 Admin"""
    u = user or session.get("user", {})
    return u.get("role") == "admin" or "manage_system" in (u.get("permissions") or [])


def _cache_only_mode(user=None):
    """PM/Investor 仅读缓存，不访问 DB"""
    u = user or session.get("user", {})
    return u.get("role") in ("project_manager", "investor")


# partner_id -> spv_id 映射，从 spv_config 派生；无 DB 时 fallback
def _get_partner_spv_map():
    try:
        from project_loader import get_partner_spv_map
        m = get_partner_spv_map()
        if m:
            return dict(m)
    except Exception:
        pass
    return {"partner_beta": "kn", "kn": "kn", "docking": "docking"}


DEFAULT_STAT_DATE = "2026-02-25"


def _get_producer_config(spv_id):
    """从 spv_config（或 producers.json fallback）获取生产商配置，支持环境变量覆盖汇率"""
    producers = load_producers()
    p = producers.get(spv_id, {})
    if not p:
        return None
    rate = p.get("exchange_rate", 1)
    currency = p.get("currency", "USD")
    # 当本币非 USD 且汇率为 1 时，尝试从 producers.json 获取正确汇率（DB 可能缺省）
    if currency != "USD" and (not rate or rate <= 1):
        try:
            with open(PRODUCERS_PATH, "r", encoding="utf-8") as f:
                fallback = json.load(f)
            fb = fallback.get("producers", {}).get(spv_id, {})
            if fb.get("exchange_rate"):
                rate = float(fb["exchange_rate"])
        except Exception:
            pass
    env_key = f"{spv_id.upper()}_EXCHANGE_RATE"
    if os.getenv(env_key):
        try:
            rate = float(os.getenv(env_key))
        except ValueError:
            pass
    return {
        "currency": currency,
        "exchange_rate": rate,
    }


def _load_risk_data_for_partner(partner_id, partner):
    """
    加载风控数据：优先从统一缓存读取（生产商页面数据一次性缓存，打开即用）
    若 partner 映射到 spv_id 且无缓存，返回空列表（用户需点击刷新从 DB 拉取）
    非生产商仍从 producers 读取
    """
    spv_id = _get_partner_spv_map().get(partner_id)
    if spv_id:
        try:
            from kn_risk_cache import load_risk_cache
            risk_data, _ = load_risk_cache(spv_id)
            if risk_data:
                return risk_data
            return []
        except Exception:
            return []
    return partner.get("risk_data", [])


def _allowed_partner_ids(user):
    """用户可访问的 partner/producer id 列表（仅读 JSON，避免 DB）"""
    allowed = list(load_partners().get("assignments", {}).get(user["username"], []))
    producers = load_producers(json_only=True)
    producer_ids = [p["id"] for p in producers.values() if str(p.get("status", "active")).lower() in ("active", "")]
    return allowed + producer_ids


def _get_partner_from_cache_or_producer(partner_id, spv_id, cache_exists, pc=None):
    """PM/Investor 有缓存时从缓存构建 partner；否则从 DB/JSON"""
    if cache_exists and pc:
        return {
            "id": spv_id,
            "name": pc.get("name", spv_id),
            "country": pc.get("region", "-"),
            "product_type": pc.get("product_type", "-"),
            "contact": pc.get("contact", "-"),
            "local_currency": pc.get("currency", "USD"),
            "exchange_rate": float(pc.get("exchange_rate", 1) or 1),
            "alerts": [],
            "priority_indicators": pc.get("priority_indicators") or {},
            "revenue_data": pc.get("revenue_data", []),
        }
    return _get_partner_or_producer(partner_id, json_only=cache_exists)


def _get_partner_or_producer(partner_id, json_only=False):
    """获取 partner 信息：优先 producers（来自 producers.json），否则从 spv_config 构建
    json_only=True: 仅从 producers.json 读取，不访问 DB，用于已有全量缓存时加速"""
    pid = str(partner_id or "").strip()
    if not pid:
        return None
    data = load_partners()
    p = data["partners"].get(pid) or data["partners"].get(pid.lower())
    if p:
        return p
    try:
        from project_loader import load_projects_with_internal_params
        if json_only:
            projects = load_projects_with_internal_params(json_only=True, skip_priority_indicators=True)
        else:
            projects = load_projects_with_internal_params()
        prod = projects.get(pid) or projects.get(pid.lower())
    except Exception:
        prod = None
    if not prod:
        prod = load_producers(json_only=json_only).get(pid) or load_producers(json_only=json_only).get(pid.lower())
    if prod and str(prod.get("status", "active")).lower() in ("active", ""):
        return {
            "id": prod["id"],
            "name": prod.get("name", prod["id"]),
            "country": prod.get("region", "-"),
            "product_type": prod.get("product_type", "-"),
            "contact": prod.get("contact", "-"),
            "local_currency": prod.get("currency", "USD"),
            "exchange_rate": prod.get("exchange_rate", 1),
            "alerts": [],
            "priority_indicators": prod.get("priority_indicators") or {},
            "revenue_data": prod.get("revenue_data", []),
        }
    return None


@app.route("/partner/<partner_id>/risk")
@login_required
def partner_risk(partner_id):
    user = _user_with_role_label(session["user"], _get_lang())
    allowed = _allowed_partner_ids(user)
    if partner_id not in allowed and user["role"] not in ("admin", "risk"):
        return redirect(url_for("dashboard"))
    spv_id, cache_exists, valid_spv = _get_spv_id_and_cache(partner_id)
    if _cache_only_mode(user) and not cache_exists:
        return redirect(url_for("partner_manage"))
    if spv_id not in valid_spv and valid_spv:
        return redirect(url_for("partner_manage"))
    pc, full_cache_updated, _ = _get_producer_data_from_full_cache(spv_id)
    partner = _get_partner_from_cache_or_producer(partner_id, spv_id, cache_exists, pc)
    if not partner:
        return redirect(url_for("partner_manage"))
    cache_last_updated = None
    if cache_exists:
        risk_data = (pc or {}).get("risk_data", [])
        if full_cache_updated:
            cache_last_updated = full_cache_updated[:19].replace("T", " ")
    elif spv_id:
        try:
            from kn_risk_cache import load_risk_cache
            risk_data, cache_last_updated = load_risk_cache(spv_id)
            risk_data = risk_data or []
            if cache_last_updated:
                cache_last_updated = cache_last_updated[:19].replace("T", " ")
        except Exception:
            risk_data = []
            cache_last_updated = None
    else:
        risk_data = _load_risk_data_for_partner(partner_id, partner)
    if cache_exists:
        local_currency = partner.get("local_currency", "USD") or "USD"
        exchange_rate = float(partner.get("exchange_rate", 1) or 1)
        env_key = f"{spv_id.upper()}_EXCHANGE_RATE"
        if os.getenv(env_key):
            try:
                exchange_rate = float(os.getenv(env_key))
            except ValueError:
                pass
    elif spv_id:
        cfg = _get_producer_config(spv_id)
        local_currency = (cfg.get("currency") if cfg else None) or partner.get("local_currency", "USD")
        exchange_rate = (cfg.get("exchange_rate") if cfg else None) or partner.get("exchange_rate", 1)
    else:
        local_currency = partner.get("local_currency", "USD")
        exchange_rate = partner.get("exchange_rate", 1)

    # 优先级指标：有缓存时用 pc.priority_indicators；PM/Investor 不访问 DB
    priority_indicators = (pc or {}).get("priority_indicators") if cache_exists else None
    if not priority_indicators:
        priority_indicators = partner.get("priority_indicators") or {}
    need_load = not (priority_indicators.get("priority_principal") and priority_indicators.get("priority_yield"))
    if spv_id and need_load and not _cache_only_mode(user):
        try:
            from spv_internal_params import load_priority_indicators_for_spv, compute_priority_from_risk_data
            pi = load_priority_indicators_for_spv(spv_id, risk_data=risk_data, exchange_rate=exchange_rate)
            if pi:
                priority_indicators = pi
            elif risk_data:
                pi = compute_priority_from_risk_data(spv_id, risk_data, exchange_rate)
                if pi:
                    priority_indicators = pi
        except Exception:
            pass

    return render_template(
        "partner_risk.html",
        user=user,
        partner=partner,
        risk_data=risk_data,
        alerts=partner.get("alerts", []),
        priority_indicators=priority_indicators,
        local_currency=local_currency,
        exchange_rate=exchange_rate,
        cache_last_updated=cache_last_updated,
        use_risk_cache=bool(spv_id),
        data_source="cache" if cache_exists else "database",
    )


def _get_spv_id_and_cache(partner_id):
    """获取 spv_id 和 cache 状态；PM/Investor 仅从缓存，Admin 可走 DB"""
    cache_only = _cache_only_mode()
    if cache_only:
        try:
            from kn_producer_cache import load_producer_full_cache
            data, _ = load_producer_full_cache()
            producers = (data or {}).get("producers", {})
            spv_map = {pid: pid for pid in producers}
            if "kn" in spv_map:
                spv_map["partner_beta"] = "kn"
            cache_exists = bool(producers)
            valid_spv = set(producers.keys()) if producers else set()
        except Exception:
            spv_map, cache_exists, valid_spv = {}, False, set()
    else:
        try:
            from project_loader import get_partner_spv_map
            spv_map = get_partner_spv_map(json_only=True)
            if not spv_map:
                spv_map = _get_partner_spv_map()
        except Exception:
            spv_map = _get_partner_spv_map()
        spv_id = spv_map.get(partner_id) or partner_id
        _, _, cache_exists = _get_producer_data_from_full_cache(spv_id)
        valid_spv = set(spv_map.values() or []) | set(load_producers(json_only=cache_exists).keys() or [])
        return spv_id, cache_exists, valid_spv
    spv_id = spv_map.get(partner_id) or partner_id
    if not cache_exists:
        _, _, cache_exists = _get_producer_data_from_full_cache(spv_id)
    return spv_id, cache_exists, valid_spv


@app.route("/partner/<partner_id>/vintage/<disbursement_month>")
@login_required
def vintage_portfolio(partner_id, disbursement_month):
    user = session["user"]
    if partner_id not in _allowed_partner_ids(user) and user["role"] not in ("admin", "risk"):
        return redirect(url_for("dashboard"))
    spv_id, cache_exists, valid_spv = _get_spv_id_and_cache(partner_id)
    partner = _get_partner_or_producer(partner_id, json_only=cache_exists)
    if not partner:
        return redirect(url_for("partner_manage"))

    stat_date = request.args.get("stat_date", "").strip()
    page = max(1, request.args.get("page", 1, type=int))
    per_page = 200

    partner_loans = []
    server_pagination = None
    if spv_id in valid_spv:
        if not stat_date:
            pc, _, cache_exists = _get_producer_data_from_full_cache(spv_id)
            if cache_exists:
                risk_data = (pc or {}).get("risk_data", [])
            else:
                risk_data = []
                try:
                    from kn_risk_cache import load_risk_cache
                    risk_data, _ = load_risk_cache(spv_id)
                except Exception:
                    pass
            if risk_data:
                latest = sorted(risk_data, key=lambda r: r.get("stat_date", ""), reverse=True)[0]
                stat_date = latest.get("stat_date", DEFAULT_STAT_DATE)
            else:
                stat_date = DEFAULT_STAT_DATE
        try:
            from kn_risk_query import query_loans_by_vintage_month
            partner_loans, total_count = query_loans_by_vintage_month(
                spv_id, stat_date, disbursement_month, page=page, per_page=per_page
            )
            total_pages = max(1, (total_count + per_page - 1) // per_page) if total_count > 0 else 1
            server_pagination = {
                "total_count": total_count,
                "page": page,
                "total_pages": total_pages,
                "per_page": per_page,
                "base_url": url_for("vintage_portfolio", partner_id=partner_id, disbursement_month=disbursement_month),
                "query_params": {"stat_date": stat_date} if stat_date else {},
            }
        except Exception:
            pass

    if not partner_loans and not server_pagination:
        portfolio_data = load_vintage_portfolio()
        partner_loans = portfolio_data.get(partner_id, {}).get(disbursement_month, [])

    stats = _portfolio_stats(partner_loans)
    return render_template(
        "portfolio_asset.html",
        user=user,
        partner=partner,
        page_title=f"Vintage {disbursement_month}",
        loans=partner_loans,
        stat_date=stat_date or "-",
        server_pagination=server_pagination,
        **stats,
    )


@app.route("/partner/<partner_id>/dpd/<bucket>")
@login_required
def dpd_portfolio(partner_id, bucket):
    user = session["user"]
    if partner_id not in _allowed_partner_ids(user) and user["role"] not in ("admin", "risk"):
        return redirect(url_for("dashboard"))
    spv_id, cache_exists, valid_spv = _get_spv_id_and_cache(partner_id)
    partner = _get_partner_or_producer(partner_id, json_only=cache_exists)
    if not partner:
        return redirect(url_for("partner_manage"))

    stat_date = request.args.get("stat_date", "").strip()
    page = max(1, request.args.get("page", 1, type=int))
    per_page = 200

    partner_loans = []
    total_count = 0
    server_pagination = None
    if spv_id in valid_spv:
        if not stat_date:
            pc, _, cache_exists = _get_producer_data_from_full_cache(spv_id)
            if cache_exists:
                risk_data = (pc or {}).get("risk_data", [])
            else:
                risk_data = []
                try:
                    from kn_risk_cache import load_risk_cache
                    risk_data, _ = load_risk_cache(spv_id)
                except Exception:
                    pass
            if risk_data:
                latest = sorted(risk_data, key=lambda r: r.get("stat_date", ""), reverse=True)[0]
                stat_date = latest.get("stat_date", DEFAULT_STAT_DATE)
            else:
                stat_date = DEFAULT_STAT_DATE
        try:
            from kn_risk_query import query_loans_by_dpd_bucket
            partner_loans, total_count = query_loans_by_dpd_bucket(spv_id, stat_date, bucket, page=page, per_page=per_page)
            total_pages = max(1, (total_count + per_page - 1) // per_page) if total_count > 0 else 1
            server_pagination = {
                "total_count": total_count,
                "page": page,
                "total_pages": total_pages,
                "per_page": per_page,
                "base_url": url_for("dpd_portfolio", partner_id=partner_id, bucket=bucket),
                "query_params": {"stat_date": stat_date} if stat_date else {},
            }
        except Exception:
            pass

    if not partner_loans and not server_pagination:
        portfolio_data = load_dpd_portfolio()
        partner_loans = portfolio_data.get(partner_id, {}).get(bucket, [])

    stats = _portfolio_stats(partner_loans)
    return render_template(
        "portfolio_asset.html",
        user=user,
        partner=partner,
        page_title=f"DPD {bucket}",
        loans=partner_loans,
        stat_date=stat_date or "-",
        server_pagination=server_pagination,
        **stats,
    )


@app.route("/partner/<partner_id>/maturity/<maturity_month>")
@login_required
def maturity_portfolio(partner_id, maturity_month):
    user = session["user"]
    if partner_id not in _allowed_partner_ids(user) and user["role"] not in ("admin", "risk"):
        return redirect(url_for("dashboard"))
    spv_id, cache_exists, valid_spv = _get_spv_id_and_cache(partner_id)
    partner = _get_partner_or_producer(partner_id, json_only=cache_exists)
    if not partner:
        return redirect(url_for("partner_manage"))

    stat_date = request.args.get("stat_date", "").strip()
    page = max(1, request.args.get("page", 1, type=int))
    per_page = 200

    partner_loans = []
    server_pagination = None
    if spv_id in valid_spv:
        if not stat_date:
            pc, _, cache_exists = _get_producer_data_from_full_cache(spv_id)
            if cache_exists:
                risk_data = (pc or {}).get("risk_data", [])
            else:
                risk_data = []
                try:
                    from kn_risk_cache import load_risk_cache
                    risk_data, _ = load_risk_cache(spv_id)
                except Exception:
                    pass
            if risk_data:
                latest = sorted(risk_data, key=lambda r: r.get("stat_date", ""), reverse=True)[0]
                stat_date = latest.get("stat_date", DEFAULT_STAT_DATE)
            else:
                stat_date = DEFAULT_STAT_DATE
        try:
            from kn_risk_query import query_loans_by_maturity_month
            partner_loans, total_count = query_loans_by_maturity_month(
                spv_id, stat_date, maturity_month, page=page, per_page=per_page
            )
            total_pages = max(1, (total_count + per_page - 1) // per_page) if total_count > 0 else 1
            server_pagination = {
                "total_count": total_count,
                "page": page,
                "total_pages": total_pages,
                "per_page": per_page,
                "base_url": url_for("maturity_portfolio", partner_id=partner_id, maturity_month=maturity_month),
                "query_params": {"stat_date": stat_date} if stat_date else {},
            }
        except Exception:
            pass

    if not partner_loans and not server_pagination:
        portfolio_data = load_maturity_portfolio()
        partner_loans = portfolio_data.get(partner_id, {}).get(maturity_month, [])

    stats = _portfolio_stats(partner_loans)
    return render_template(
        "portfolio_asset.html",
        user=user,
        partner=partner,
        page_title=f"到期月 {maturity_month}",
        loans=partner_loans,
        stat_date=stat_date or "-",
        server_pagination=server_pagination,
        **stats,
    )


@app.route("/partner/<partner_id>/loan/<loan_id>")
@login_required
def loan_detail(partner_id, loan_id):
    user = session["user"]
    if partner_id not in _allowed_partner_ids(user) and user["role"] not in ("admin", "risk"):
        return redirect(url_for("dashboard"))
    spv_id, cache_exists, valid_spv = _get_spv_id_and_cache(partner_id)
    partner = _get_partner_or_producer(partner_id, json_only=cache_exists)
    if not partner:
        return redirect(url_for("partner_manage"))
    loan = None
    schedule = []
    repayments = []
    customer_info = {}
    contract_no = "-"

    # 优先从数据库查询（KN 等有 spv_id 的生产商），仿照查询页面内容
    if spv_id in valid_spv:
        try:
            from risk_query import query_loan_detail
            result = query_loan_detail(loan_id, spv_id=spv_id)
            if not result.get("error") and result.get("loans"):
                first = result["loans"][0]
                contract_no = result.get("contract_no", "-")
                status = first.get("status", {})
                schedule = first.get("schedule", [])
                records = first.get("records", [])

                # 合并 raw_loan + calc_overdue 信息
                from kn_risk_query import get_loan_overdue_info, get_customer_info
                overdue = get_loan_overdue_info(loan_id, spv_id)
                customer_info = get_customer_info(status.get("customer_id"))

                loan = {
                    "loan_id": status.get("loan_id", loan_id),
                    "disbursement_amount": float(status.get("disbursement_amount") or 0),
                    "disbursement_time": status.get("disbursement_time"),
                    "term_month": status.get("term_months"),
                    "loan_maturity_date": status.get("loan_maturity_date"),
                    "customer_id": status.get("customer_id"),
                    "contract_no": status.get("contract_no"),
                    "spv_id": status.get("spv_id"),
                    "loan_status": "active" if overdue.get("loan_status") == 1 else "overdue" if overdue.get("loan_status") == 2 else "closed",
                    "dpd": overdue.get("dpd", 0),
                    "outstanding_principal": overdue.get("outstanding_principal", 0),
                }

                # 还款计划：标记已还期次
                paid_terms = {r.get("repayment_term") for r in records if r.get("repayment_term") and r.get("repayment_term") > 0}
                for s in schedule:
                    s["period"] = s.get("period_no", s.get("period"))
                    s["status"] = "paid" if s.get("period_no") in paid_terms else "pending"

                # 还款信息：映射字段名
                repayments = []
                for r in records:
                    repayments.append({
                        "repayment_type": r.get("repayment_type"),
                        "repayment_date": r.get("repayment_date"),
                        "repayment_term": r.get("repayment_term"),
                        "total_repayment": r.get("total_repayment"),
                        "principal_repayment": r.get("principal_repayment"),
                        "interest_repayment": r.get("interest_repayment"),
                        "penalty_repayment": r.get("penalty_repayment"),
                        "extension_fee": r.get("extension_fee"),
                        "waiver_amount": r.get("waiver_amount"),
                        "repayment_txn_id": r.get("repayment_txn_id"),
                        "is_settled": r.get("is_settled"),
                    })
        except Exception:
            pass

    if not loan:
        loan = _find_loan(partner_id, loan_id)
        if not loan:
            return redirect(url_for("partner_risk", partner_id=partner_id))
        details = load_loan_details().get(loan_id, {})
        schedule = details.get("schedule", [])
        repayments = details.get("repayments", [])
        customer_info = details.get("customer_info", {})
        contract_no = "-"

    return render_template(
        "loan_detail.html",
        user=user,
        partner=partner,
        loan=loan,
        schedule=schedule,
        repayments=repayments,
        customer_info=customer_info,
        contract_no=contract_no,
    )


@app.route("/partner/<partner_id>/cashflow")
@login_required
def partner_cashflow(partner_id):
    user = _user_with_role_label(session["user"], _get_lang())
    pid = str(partner_id or "").strip()
    spv_id, cache_exists, valid_spv = _get_spv_id_and_cache(pid)
    if _cache_only_mode(user) and not cache_exists:
        return redirect(url_for("partner_manage"))
    allowed = _allowed_partner_ids(user)
    allowed_lower = {str(x).lower() for x in allowed}
    producer_ids = list(valid_spv) if cache_exists else [str(p["id"]).lower() for p in load_producers(json_only=True).values()]
    has_manage = "manage_partners" in user.get("permissions", [])
    can_access = (
        pid in allowed
        or pid.lower() in allowed_lower
        or user["role"] in ("admin", "risk")
        or (has_manage and pid.lower() in producer_ids)
    )
    if not can_access:
        return redirect(url_for("dashboard"))
    pc, full_cache_updated, _ = _get_producer_data_from_full_cache(spv_id)
    partner = _get_partner_from_cache_or_producer(pid, spv_id, cache_exists, pc)
    if not partner:
        return redirect(url_for("partner_manage"))
    if cache_exists:
        local_currency = (pc or {}).get("currency") or partner.get("local_currency", "USD") or "USD"
        exchange_rate = float((pc or {}).get("exchange_rate", 1) or partner.get("exchange_rate", 1) or 1)
        env_key = f"{spv_id.upper()}_EXCHANGE_RATE"
        if os.getenv(env_key):
            try:
                exchange_rate = float(os.getenv(env_key))
            except ValueError:
                pass
    else:
        cfg = _get_producer_config(spv_id)
        local_currency = (cfg.get("currency") if cfg else None) or partner.get("local_currency", "USD")
        exchange_rate = float((cfg.get("exchange_rate") if cfg else None) or partner.get("exchange_rate", 1) or 1)
    cashflow_data = []
    cache_last_updated = None
    use_cashflow_cache = bool(spv_id)
    collection_rate = 0.98
    if cache_exists:
        cashflow_data = (pc or {}).get("cashflow_data", [])
        if full_cache_updated:
            cache_last_updated = full_cache_updated[:19].replace("T", " ")
        rev_list = (pc or {}).get("revenue_data", [])
        if rev_list:
            cr = rev_list[-1].get("collection_rate", 0.98) or 0.98
            if cr >= 0.5:
                collection_rate = cr
            elif len(rev_list) >= 2:
                collection_rate = rev_list[-2].get("collection_rate", 0.98) or 0.98
    else:
        try:
            from kn_cashflow_cache import load_cashflow_cache
            cached_cf, cache_last_updated, cr = load_cashflow_cache(spv_id)
            if cached_cf:
                cashflow_data = cached_cf
                rev_data = partner.get("revenue_data", [])
                if rev_data:
                    cr = rev_data[-1].get("collection_rate", 0.98) or 0.98
                    collection_rate = cr if cr >= 0.5 else (rev_data[-2].get("collection_rate", 0.98) or 0.98 if len(rev_data) >= 2 else 0.98)
                else:
                    collection_rate = cr if cr >= 0.5 else 0.98
            else:
                from kn_cashflow import compute_cashflow_forecast
                rev_data = partner.get("revenue_data", [])
                if rev_data:
                    cr = rev_data[-1].get("collection_rate", 0.98) or 0.98
                    collection_rate = cr if cr >= 0.5 else (rev_data[-2].get("collection_rate", 0.98) or 0.98 if len(rev_data) >= 2 else 0.98)
                cf = compute_cashflow_forecast(spv_id=spv_id, months_ahead=12, collection_rate=collection_rate)
                cashflow_data = cf.get("forecast", [])
        except Exception:
            pass
    # 使用正确的回收率重新计算预期回收（避免缓存中使用了错误回收率的历史数据）
    for row in cashflow_data:
        p = float(row.get("principal") or 0)
        i = float(row.get("interest") or 0)
        row["expected_inflow"] = int(round((p + i) * collection_rate))
    if cache_last_updated:
        cache_last_updated = cache_last_updated[:19].replace("T", " ")
    return render_template(
        "partner_cashflow.html",
        user=user,
        partner=partner,
        cashflow_data=cashflow_data,
        collection_rate=collection_rate,
        local_currency=local_currency,
        exchange_rate=exchange_rate,
        use_cashflow_cache=use_cashflow_cache,
        cache_last_updated=cache_last_updated,
        data_source="cache" if cache_exists else "database",
    )


@app.route("/partner/<partner_id>/revenue")
@login_required
def partner_revenue(partner_id):
    user = _user_with_role_label(session["user"], _get_lang())
    pid = str(partner_id or "").strip()
    spv_id, cache_exists, valid_spv = _get_spv_id_and_cache(pid)
    if _cache_only_mode(user) and not cache_exists:
        return redirect(url_for("partner_manage"))
    allowed = _allowed_partner_ids(user)
    allowed_lower = {str(x).lower() for x in allowed}
    producer_ids = list(valid_spv) if cache_exists else [str(p["id"]).lower() for p in load_producers(json_only=True).values()]
    has_manage = "manage_partners" in user.get("permissions", [])
    can_access = (
        pid in allowed
        or pid.lower() in allowed_lower
        or user["role"] in ("admin", "risk")
        or (has_manage and pid.lower() in producer_ids)
    )
    if not can_access:
        return redirect(url_for("dashboard"))
    pc, full_cache_updated, _ = _get_producer_data_from_full_cache(spv_id)
    partner = _get_partner_from_cache_or_producer(pid, spv_id, cache_exists, pc)
    if not partner:
        return redirect(url_for("partner_manage"))
    cache_last_updated = None
    use_revenue_cache = bool(spv_id)
    if cache_exists:
        revenue_data = (pc or {}).get("revenue_data", [])
        if full_cache_updated:
            cache_last_updated = full_cache_updated[:19].replace("T", " ")
    else:
        revenue_data = partner.get("revenue_data", [])
        if not revenue_data:
            producers = load_producers()
            rev_prod = producers.get(pid) or producers.get(pid.lower())
            if rev_prod:
                revenue_data = rev_prod.get("revenue_data", [])
        if spv_id:
            try:
                from kn_revenue_cache import load_revenue_cache
                cached_rev, cache_last_updated = load_revenue_cache(spv_id)
                if cached_rev:
                    revenue_data = cached_rev
            except Exception:
                pass
    if cache_exists:
        local_currency = partner.get("local_currency", "USD") or "USD"
        exchange_rate = float(partner.get("exchange_rate", 1) or 1)
        env_key = f"{spv_id.upper()}_EXCHANGE_RATE"
        if os.getenv(env_key):
            try:
                exchange_rate = float(os.getenv(env_key))
            except ValueError:
                pass
    elif spv_id:
        cfg = _get_producer_config(spv_id)
        local_currency = (cfg.get("currency") if cfg else None) or partner.get("local_currency", "USD")
        exchange_rate = (cfg.get("exchange_rate") if cfg else None) or partner.get("exchange_rate", 1)
    else:
        local_currency = partner.get("local_currency", "USD")
        exchange_rate = partner.get("exchange_rate", 1)
    if cache_last_updated:
        cache_last_updated = cache_last_updated[:19].replace("T", " ")
    return render_template(
        "partner_revenue.html",
        user=user,
        partner=partner,
        revenue_data=revenue_data,
        local_currency=local_currency,
        data_source="cache" if cache_exists else "database",
        exchange_rate=exchange_rate,
        use_revenue_cache=use_revenue_cache,
        cache_last_updated=cache_last_updated,
    )


@app.route("/transaction/apply")
@login_required
def transaction_apply():
    user = session["user"]
    if "apply_transactions" not in user.get("permissions", []):
        return redirect(url_for("dashboard"))
    txn_data = load_transactions()
    apps = [a for a in txn_data["applications"] if a["applicant"] == user["username"]]
    partner_data = load_partners()
    partner_ids = partner_data["assignments"].get(user["username"], [])
    partners_list = [
        {"id": pid, "name": partner_data["partners"][pid]["name"]}
        for pid in partner_ids if pid in partner_data["partners"]
    ]
    return render_template(
        "transaction_apply.html",
        user=user,
        applications=apps,
        partners=partners_list,
    )


@app.route("/api/db/status")
@login_required
def api_db_status():
    """数据库连接状态检查（供数据查询等页面使用）"""
    try:
        from db_connect import get_connection
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("SELECT 1")
        cur.fetchone()
        cur.close()
        conn.close()
        return jsonify({"ok": True, "message": "数据库连接正常"})
    except Exception as e:
        err = str(e)
        hint = ""
        if "translate host name" in err.lower() or "nodename" in err.lower():
            hint = "域名解析失败，请检查：1) 网络是否正常；2) 是否需连接 VPN；3) 在系统终端（非 IDE）运行 python3 app.py"
        elif "timeout" in err.lower() or "timed out" in err.lower():
            hint = "连接超时，请检查：1) RDS 白名单是否包含当前 IP；2) 是否需通过 VPN/内网访问；3) 公网地址是否已开启"
        elif "connection refused" in err.lower():
            hint = "连接被拒绝，请检查端口、防火墙及 RDS 是否运行"
        elif "password" in err.lower() or "authentication" in err.lower():
            hint = "认证失败，请检查 .env 中的 DB_USER 和 DB_PASSWORD"
        return jsonify({"ok": False, "message": err, "hint": hint})


@app.route("/risk/query")
@login_required
def risk_query_page():
    user = session["user"]
    if "data_query" not in user.get("permissions", []):
        return redirect(url_for("dashboard"))
    return render_template("risk_query.html", user=user)


@app.route("/api/risk/query/loan/<loan_id>")
@login_required
def api_risk_query_loan(loan_id):
    if "data_query" not in session["user"].get("permissions", []):
        return jsonify({"error": "权限不足"}), 403
    from risk_query import query_loan_detail
    result = query_loan_detail(loan_id)
    return jsonify(result)


@app.route("/api/partner/refresh-all-cache", methods=["POST"])
@login_required
def api_refresh_all_producer_cache():
    """Admin 启动后台刷新：立即返回，刷新在后台执行，可通过 /api/partner/refresh-status 轮询"""
    if not _is_admin():
        return jsonify({"error": "权限不足"}), 403
    try:
        from kn_producer_cache import refresh_producer_full_cache_async
        refresh_producer_full_cache_async()
        return jsonify({"started": True})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/partner/refresh-status")
@login_required
def api_refresh_status():
    """轮询刷新状态与日志（Admin）"""
    if not _is_admin():
        return jsonify({"error": "权限不足"}), 403
    try:
        from kn_producer_cache import get_refresh_status, load_refresh_log, load_cache_meta
        st = get_refresh_status()
        logs = load_refresh_log()
        meta = load_cache_meta()
        logs_text = "".join(logs) if isinstance(logs, list) else (logs or "")
        out = {
            "running": st.get("running", False),
            "logs": logs_text,
            "last_updated": meta.get("last_updated") if meta else None,
            "system_cutover_date": meta.get("system_cutover_date") if meta else None,
        }
        if st.get("result"):
            r = st["result"]
            if "error" in r:
                out["error"] = r["error"]
            if r.get("ok"):
                out["last_updated"] = r.get("last_updated")
                out["system_cutover_date"] = r.get("system_cutover_date")
                out["producer_count"] = r.get("producer_count")
        return jsonify(out)
    except Exception as e:
        return jsonify({"error": str(e), "running": False}), 500


@app.route("/api/partner/<partner_id>/refresh-risk", methods=["POST"])
@login_required
def api_refresh_risk(partner_id):
    """刷新生产商风控数据缓存：从数据库重新计算核心指标、DPD、Vintage 等并保存（仅 Admin）"""
    if not _is_admin():
        return jsonify({"error": "权限不足"}), 403
    spv_map = _get_partner_spv_map()
    spv_id = spv_map.get(partner_id) or spv_map.get(str(partner_id).lower()) or partner_id
    spv_id_lower = str(spv_id).strip().lower()
    valid_spv = set(spv_map.values() or []) | set(load_producers().keys() or [])
    valid_spv_lower = {str(v).strip().lower() for v in valid_spv}
    if spv_id_lower not in valid_spv_lower:
        return jsonify({"error": f"未知生产商: {partner_id}"}), 400
    cfg = _get_producer_config(spv_id_lower)
    exchange_rate = (cfg.get("exchange_rate") if cfg else 1) or 1
    currency = (cfg.get("currency") if cfg else "USD") or "USD"
    try:
        from kn_risk_cache import refresh_risk_cache
        from kn_producer_cache import update_producer_risk_in_full_cache
        result = refresh_risk_cache(spv_id_lower, exchange_rate, currency)
        if "error" in result:
            return jsonify(result), 500
        update_producer_risk_in_full_cache(spv_id_lower, exchange_rate, currency)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/partner/<partner_id>/refresh-revenue", methods=["POST"])
@login_required
def api_refresh_revenue(partner_id):
    """刷新生产商收益数据缓存：从数据库重新计算并保存（仅 Admin）"""
    if not _is_admin():
        return jsonify({"error": "权限不足"}), 403
    spv_id = _get_partner_spv_map().get(partner_id) or partner_id
    valid_spv = set(_get_partner_spv_map().values() or []) | set(load_producers().keys() or [])
    if spv_id not in valid_spv:
        return jsonify({"error": f"未知生产商: {partner_id}"}), 400
    cfg = _get_producer_config(spv_id)
    exchange_rate = float((cfg.get("exchange_rate") if cfg else 1) or 1)
    currency = (cfg.get("currency") if cfg else "USD") or "USD"
    try:
        from kn_revenue_cache import refresh_revenue_cache
        from kn_producer_cache import update_producer_revenue_in_full_cache
        result = refresh_revenue_cache(spv_id, exchange_rate, currency)
        if "error" in result:
            return jsonify(result), 500
        update_producer_revenue_in_full_cache(spv_id)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/partner/<partner_id>/refresh-cashflow", methods=["POST"])
@login_required
def api_refresh_cashflow(partner_id):
    """刷新生产商现金流数据缓存：从数据库重新计算并保存（仅 Admin）"""
    if not _is_admin():
        return jsonify({"error": "权限不足"}), 403
    spv_id = _get_partner_spv_map().get(partner_id) or partner_id
    valid_spv = set(_get_partner_spv_map().values() or []) | set(load_producers().keys() or [])
    if spv_id not in valid_spv:
        return jsonify({"error": f"未知生产商: {partner_id}"}), 400
    cfg = _get_producer_config(spv_id)
    exchange_rate = float((cfg.get("exchange_rate") if cfg else 1) or 1)
    currency = (cfg.get("currency") if cfg else "USD") or "USD"
    coll_rate = 0.98
    try:
        from kn_revenue_cache import load_revenue_cache
        cached_rev, _ = load_revenue_cache(spv_id)
        if cached_rev:
            cr = cached_rev[-1].get("collection_rate", 0.98) or 0.98
            coll_rate = cr if cr >= 0.5 else (cached_rev[-2].get("collection_rate", 0.98) or 0.98 if len(cached_rev) >= 2 else 0.98)
    except Exception:
        pass
    try:
        from kn_cashflow_cache import refresh_cashflow_cache
        from kn_producer_cache import update_producer_cashflow_in_full_cache
        result = refresh_cashflow_cache(spv_id, exchange_rate, currency, coll_rate)
        if "error" in result:
            return jsonify(result), 500
        update_producer_cashflow_in_full_cache(spv_id)
        return jsonify(result)
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/risk/query/disbursements", methods=["POST"])
@login_required
def api_risk_query_disbursements():
    if "data_query" not in session["user"].get("permissions", []):
        return jsonify({"error": "权限不足"}), 403
    data = request.get_json() or {}
    query_date = data.get("date", "").strip()
    from risk_query import query_daily_disbursements
    result = query_daily_disbursements(query_date)
    return jsonify(result)


@app.route("/transaction/review")
@login_required
def transaction_review():
    user = session["user"]
    if "approve_transactions" not in user.get("permissions", []):
        return redirect(url_for("dashboard"))
    txn_data = load_transactions()
    return render_template(
        "transaction_review.html",
        user=user,
        applications=txn_data["applications"],
    )


@app.route("/api/dd/upload", methods=["POST"])
@login_required
def dd_upload():
    if "file" not in request.files:
        return jsonify({"success": False, "message": "未选择文件"}), 400
    f = request.files["file"]
    item_id = request.form.get("item_id", "unknown")
    if f.filename:
        filename = secure_filename(f"{item_id}_{f.filename}")
        user_dir = os.path.join(UPLOAD_DIR, session["user"]["username"])
        os.makedirs(user_dir, exist_ok=True)
        f.save(os.path.join(user_dir, filename))
        return jsonify({"success": True, "filename": filename})
    return jsonify({"success": False, "message": "文件无效"}), 400


@app.route("/api/dd/template/<module_id>")
@login_required
def dd_template_download(module_id):
    template_file = f"{module_id}_checklist.xlsx"
    template_path = os.path.join(DD_TEMPLATES_DIR, template_file)
    if os.path.exists(template_path):
        return send_from_directory(DD_TEMPLATES_DIR, template_file, as_attachment=True)
    dd = load_dd_checklist()
    mod = next((m for m in dd["modules"] if m["id"] == module_id), None)
    if not mod:
        return jsonify({"error": "模块不存在"}), 404
    content = f"{mod['name']} ({mod['name_en']}) - 尽调清单\n{'='*50}\n\n"
    for i, item in enumerate(mod["items"], 1):
        req = "【必填】" if item["required"] else "【选填】"
        content += f"{i}. {req} {item['name']}\n   文件名：\n   备注：\n\n"
    from flask import Response
    return Response(
        content,
        mimetype="text/plain; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={module_id}_checklist.txt"}
    )


def _json_to_excel_rows(data_list):
    """将 JSON 列表转为 Excel 行：标量直接写入，dict/list 转为 JSON 字符串"""
    if not data_list:
        return [], []
    all_keys = set()
    for row in data_list:
        if isinstance(row, dict):
            all_keys.update(row.keys())
    headers = sorted(all_keys)
    rows = []
    for row in data_list:
        if not isinstance(row, dict):
            rows.append([str(row)])
            continue
        r = []
        for h in headers:
            v = row.get(h)
            if isinstance(v, (dict, list)):
                r.append(json.dumps(v, ensure_ascii=False) if v is not None else "")
            else:
                r.append(v if v is not None else "")
        rows.append(r)
    return headers, rows


def _flatten_priority_indicators(pi):
    """将 priority_indicators 展平为单行 dict，供 Excel 导出"""
    if not pi:
        return {}
    flat = {}
    if pi.get("priority_principal") is not None:
        flat["priority_principal"] = pi["priority_principal"]
    if pi.get("priority_yield"):
        py = pi["priority_yield"]
        flat["priority_yield_current"] = py.get("current")
        flat["priority_yield_target"] = py.get("target")
    if pi.get("leverage_ratio"):
        lr = pi["leverage_ratio"]
        flat["leverage_ratio_current"] = lr.get("current")
        flat["leverage_ratio_limit"] = lr.get("limit")
    if pi.get("coverage_ratio"):
        cov = pi["coverage_ratio"]
        flat["coverage_ratio_current"] = cov.get("current")
        flat["coverage_liquidation_line"] = cov.get("liquidation")
        flat["coverage_margin_call_line"] = cov.get("margin_call")
        flat["coverage_baseline"] = cov.get("baseline")
        b = cov.get("breakdown") or {}
        for k, v in b.items():
            flat["breakdown_" + k] = v
    if pi.get("coverage_ratio_abs"):
        cov = pi["coverage_ratio_abs"]
        flat["coverage_ratio_abs_current"] = cov.get("current")
        b = cov.get("breakdown") or {}
        for k, v in b.items():
            flat["breakdown_abs_" + k] = v
    return flat


@app.route("/api/partner/<partner_id>/download-excel")
@login_required
def api_partner_download_excel(partner_id):
    """下载生产商缓存数据为 Excel：风控、收益、现金流 三个 Tab"""
    spv_id = _get_partner_spv_map().get(partner_id) or partner_id
    valid_spv = set(_get_partner_spv_map().values() or []) | set(load_producers(json_only=True).keys() or [])
    if spv_id not in valid_spv:
        return jsonify({"error": "未知生产商"}), 404

    pc, _, cache_exists = _get_producer_data_from_full_cache(spv_id)
    risk_data = []
    revenue_data = []
    cashflow_data = []
    priority_indicators = None

    if cache_exists and pc:
        risk_data = pc.get("risk_data", [])
        revenue_data = pc.get("revenue_data", [])
        cashflow_data = pc.get("cashflow_data", [])
        priority_indicators = pc.get("priority_indicators")
        if not priority_indicators and risk_data:
            try:
                from spv_internal_params import load_priority_indicators_for_spv, compute_priority_from_risk_data
                rate = float((pc.get("exchange_rate") or 1) or 1)
                pi = load_priority_indicators_for_spv(spv_id, risk_data=risk_data, exchange_rate=rate)
                priority_indicators = pi or compute_priority_from_risk_data(spv_id, risk_data, rate)
            except Exception:
                pass
    else:
        try:
            from kn_risk_cache import load_risk_cache
            rd, _ = load_risk_cache(spv_id)
            risk_data = rd or []
        except Exception:
            pass
        try:
            from kn_revenue_cache import load_revenue_cache
            rev, _ = load_revenue_cache(spv_id)
            revenue_data = rev or []
        except Exception:
            pass
        try:
            from kn_cashflow_cache import load_cashflow_cache
            cf, _ = load_cashflow_cache(spv_id)
            cashflow_data = cf or []
        except Exception:
            pass
        if risk_data:
            try:
                from spv_internal_params import load_priority_indicators_for_spv, compute_priority_from_risk_data
                cfg = _get_producer_config(spv_id)
                rate = float((cfg.get("exchange_rate") if cfg else 1) or 1)
                pi = load_priority_indicators_for_spv(spv_id, risk_data=risk_data, exchange_rate=rate)
                priority_indicators = pi or compute_priority_from_risk_data(spv_id, risk_data, rate)
            except Exception:
                pass

    try:
        from openpyxl import Workbook
        from openpyxl.utils import get_column_letter
        from io import BytesIO

        wb = Workbook()
        wb.remove(wb.active)

        # Tab 1: 风控
        ws1 = wb.create_sheet("风控", 0)
        h1, r1 = _json_to_excel_rows(risk_data)
        if h1:
            ws1.append(h1)
            for row in r1:
                ws1.append(row)

        # Tab 2: 收益
        ws2 = wb.create_sheet("收益", 1)
        h2, r2 = _json_to_excel_rows(revenue_data)
        if h2:
            ws2.append(h2)
            for row in r2:
                ws2.append(row)

        # Tab 3: 现金流（补充回收率到每行，与 partner_cashflow 页面一致）
        ws3 = wb.create_sheet("现金流", 2)
        collection_rate = 0.98
        if revenue_data:
            cr = revenue_data[-1].get("collection_rate", 0.98) or 0.98
            collection_rate = cr if cr >= 0.5 else (revenue_data[-2].get("collection_rate", 0.98) or 0.98 if len(revenue_data) >= 2 else 0.98)
        else:
            try:
                from kn_cashflow_cache import load_cashflow_cache
                _, _, cr = load_cashflow_cache(spv_id)
                collection_rate = cr if cr and cr >= 0.5 else 0.98
            except Exception:
                pass
        cf_for_excel = []
        for row in cashflow_data:
            r = dict(row) if isinstance(row, dict) else {}
            r["collection_rate"] = round(collection_rate * 100, 1)  # 以百分比形式存储，如 98.0
            cf_for_excel.append(r)
        h3, r3 = _json_to_excel_rows(cf_for_excel)
        if h3:
            ws3.append(h3)
            for row in r3:
                ws3.append(row)

        # Tab 4: 优先级指标（含覆盖倍数拆解）
        flat_pi = _flatten_priority_indicators(priority_indicators)
        if flat_pi:
            ws4 = wb.create_sheet("优先级指标", 3)
            # 固定列顺序：主要指标在前，便于查看
            priority_order = [
                "priority_principal", "priority_yield_current", "priority_yield_target",
                "leverage_ratio_current", "leverage_ratio_limit",
                "coverage_ratio_current", "coverage_liquidation_line", "coverage_margin_call_line", "coverage_baseline",
                "breakdown_stat_date", "breakdown_m0_balance", "breakdown_m0_accrued_interest",
                "breakdown_early_repayment_overdue_discount", "breakdown_m0_interest_discounted",
                "breakdown_vtg30_predicted_default_rate", "breakdown_after_default_local", "breakdown_cash",
                "breakdown_value_local", "breakdown_exchange_rate", "breakdown_value_usd",
                "breakdown_coop_principal", "breakdown_unallocated", "breakdown_loan_usd", "breakdown_coverage_ratio",
            ]
            h4 = [k for k in priority_order if k in flat_pi]
            h4 += sorted(k for k in flat_pi if k not in h4)
            r4 = [[flat_pi.get(h) for h in h4]]
            if h4:
                ws4.append(h4)
                for row in r4:
                    ws4.append(row)

        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        prod = load_producers(json_only=True).get(spv_id, {})
        name = (prod.get("name") or prod.get("id") or spv_id).replace("/", "_").replace("\\", "_")
        filename = f"{name}_数据导出.xlsx"

        return send_file(
            buf,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            as_attachment=True,
            download_name=filename,
        )
    except ImportError:
        return jsonify({"error": "openpyxl 未安装，请运行 pip install openpyxl"}), 500
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/user")
@login_required
def api_user():
    return jsonify(session["user"])


@app.route("/logout")
def logout():
    session.pop("user", None)
    return redirect(url_for("login_page"))


if __name__ == "__main__":
    port = int(os.getenv("PORT", "5001"))
    app.run(host="0.0.0.0", port=port, debug=True)
